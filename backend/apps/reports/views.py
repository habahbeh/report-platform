"""
Report and Project views.
"""

from django.utils import timezone
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, permissions, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response as DRFResponse
from rest_framework.parsers import MultiPartParser, FormParser

from apps.templates_app.models import Item
from .models import (
    Project, Contributor, Response, TableData, GeneratedReport,
    ItemStructure, GeneratedContent
)
from .serializers import (
    ProjectSerializer, ProjectDetailSerializer, ProjectCreateSerializer,
    ProjectStatsSerializer, AggregatedDataSerializer,
    ContributorSerializer, ContributorDetailSerializer, ContributorFormSerializer,
    ResponseSerializer, ResponseCreateSerializer,
    TableDataSerializer, GeneratedReportSerializer,
    ItemStructureSerializer, ItemStructureUpdateSerializer,
    ItemStructureCreateFromTemplateSerializer,
    GeneratedContentSerializer, GeneratedContentEditSerializer,
    GeneratedContentRegenerateSerializer,
    SkeletonBuildRequestSerializer, TextGenerateRequestSerializer,
)


# ============================================
# Project ViewSet (New System)
# ============================================

class ProjectViewSet(viewsets.ModelViewSet):
    """API for Projects"""
    permission_classes = [permissions.AllowAny]  # For demo

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ProjectDetailSerializer
        if self.action == 'create':
            return ProjectCreateSerializer
        return ProjectSerializer

    def get_queryset(self):
        queryset = Project.objects.all()

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by organization
        org_id = self.request.query_params.get('organization')
        if org_id:
            queryset = queryset.filter(organization_id=org_id)

        # Filter by template
        template_id = self.request.query_params.get('template')
        if template_id:
            queryset = queryset.filter(template_id=template_id)

        return queryset.select_related('template', 'organization')

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        """Get project statistics"""
        project = self.get_object()

        total_items = Item.objects.filter(axis__template=project.template).count()
        completed_items = project.responses.values('item').distinct().count()

        contributors = project.contributors.all()

        data = {
            'total_items': total_items,
            'completed_items': completed_items,
            'items_progress': project.items_progress,

            'total_contributors': contributors.count(),
            'contributors_completed': contributors.filter(status='completed').count(),
            'contributors_in_progress': contributors.filter(status='in_progress').count(),
            'contributors_pending': contributors.filter(status__in=['pending', 'invited']).count(),

            'deadline': project.deadline,
            'days_remaining': project.days_remaining,
            'status': project.status,
        }

        return DRFResponse(data)

    @action(detail=True, methods=['get'])
    def contributors(self, request, pk=None):
        """List project contributors"""
        project = self.get_object()
        contributors = project.contributors.all()
        serializer = ContributorSerializer(contributors, many=True)
        return DRFResponse(serializer.data)

    @action(detail=True, methods=['post'])
    def add_contributor(self, request, pk=None):
        """Add a contributor to the project"""
        project = self.get_object()

        serializer = ContributorSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(project=project)
            return DRFResponse(serializer.data, status=status.HTTP_201_CREATED)
        return DRFResponse(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def invite(self, request, pk=None):
        """Send invitations to contributors"""
        project = self.get_object()

        contributor_ids = request.data.get('contributor_ids', 'all')

        if contributor_ids == 'all':
            contributors = project.contributors.filter(status='pending')
        else:
            contributors = project.contributors.filter(id__in=contributor_ids)

        # Update status and sent time
        now = timezone.now()
        for contributor in contributors:
            contributor.status = 'invited'
            contributor.invite_sent_at = now
            contributor.save(update_fields=['status', 'invite_sent_at'])

        # TODO: Actually send emails

        return DRFResponse({
            'status': 'success',
            'message': f'تم إرسال {contributors.count()} دعوة',
            'invited_count': contributors.count()
        })

    @action(detail=True, methods=['post'])
    def remind(self, request, pk=None):
        """Send reminders to incomplete contributors"""
        project = self.get_object()

        contributor_ids = request.data.get('contributor_ids', 'incomplete')

        if contributor_ids == 'incomplete':
            contributors = project.contributors.exclude(status='completed')
        else:
            contributors = project.contributors.filter(id__in=contributor_ids)

        now = timezone.now()
        for contributor in contributors:
            contributor.last_reminder_at = now
            contributor.reminder_count += 1
            contributor.save(update_fields=['last_reminder_at', 'reminder_count'])

        # TODO: Actually send reminder emails

        return DRFResponse({
            'status': 'success',
            'message': f'تم إرسال {contributors.count()} تذكير'
        })

    @action(detail=True, methods=['post'])
    def generate(self, request, pk=None):
        """Generate the final report using FullReportGenerator (branded)"""
        import threading
        import os
        from pathlib import Path
        from django.conf import settings
        from django.core.files.base import ContentFile
        from apps.export.full_report_generator import FullReportGenerator

        project = self.get_object()

        format_type = request.data.get('format', 'docx')
        options = request.data.get('options', {})

        # Optional: generate only specific axis or item
        axis_code = request.data.get('axis_code')
        item_code = request.data.get('item_code')
        axis_codes = [axis_code] if axis_code else None
        item_codes = [item_code] if item_code else None

        # Create GeneratedReport record
        generated_report = GeneratedReport.objects.create(
            project=project,
            format=format_type,
            options={**options, 'axis_code': axis_code, 'item_code': item_code},
            status='processing',
            current_step='بدء التوليد',
            created_by=request.user if request.user.is_authenticated else None
        )

        # Only transition to 'generating' for full report exports
        if not axis_code and not item_code:
            project.status = 'generating'
            project.save(update_fields=['status'])

        def do_generate():
            try:
                # Generate using FullReportGenerator (branded + project-aware)
                output_dir = Path(settings.MEDIA_ROOT) / 'generated' / str(project.id)
                output_dir.mkdir(parents=True, exist_ok=True)

                generator = FullReportGenerator(
                    project=project,
                    output_dir=str(output_dir)
                )
                results = generator.generate(
                    formats=format_type,
                    axis_codes=axis_codes,
                    item_codes=item_codes,
                )

                # Get the generated file path
                file_path = results.get(format_type)
                if not file_path or not os.path.exists(file_path):
                    generated_report.status = 'failed'
                    generated_report.error_message = f'فشل توليد ملف {format_type}'
                    generated_report.save()
                    return

                # Save file to GeneratedReport
                with open(file_path, 'rb') as f:
                    file_content = f.read()
                filename = os.path.basename(file_path)
                generated_report.file.save(filename, ContentFile(file_content))
                generated_report.file_size = len(file_content)
                generated_report.status = 'completed'
                generated_report.completed_at = timezone.now()
                generated_report.save()

                # Only publish project for full report exports
                if not axis_code and not item_code:
                    project.status = 'published'
                    project.published_at = timezone.now()
                    project.save(update_fields=['status', 'published_at'])

            except Exception as e:
                import traceback
                generated_report.status = 'failed'
                generated_report.error_message = f'{str(e)}\n{traceback.format_exc()}'
                generated_report.save()

        # Run in background thread (for demo - use Celery in production)
        thread = threading.Thread(target=do_generate)
        thread.start()

        return DRFResponse({
            'status': 'started',
            'report_id': str(generated_report.id),
            'message': 'بدأ توليد التقرير'
        })

    @action(detail=True, methods=['get'])
    def reports(self, request, pk=None):
        """List generated reports"""
        project = self.get_object()
        reports = project.generated_reports.all()
        serializer = GeneratedReportSerializer(reports, many=True)
        return DRFResponse(serializer.data)


# ============================================
# Contributor ViewSet
# ============================================

class ContributorViewSet(viewsets.ModelViewSet):
    """API for Contributors"""
    serializer_class = ContributorSerializer
    permission_classes = [permissions.AllowAny]  # For demo

    def get_queryset(self):
        return Contributor.objects.all()

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ContributorDetailSerializer
        return ContributorSerializer

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve contributor's submission"""
        contributor = self.get_object()
        contributor.status = 'completed'
        contributor.save(update_fields=['status'])
        return DRFResponse({'status': 'approved'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject contributor's submission"""
        contributor = self.get_object()
        contributor.status = 'rejected'
        contributor.rejection_reason = request.data.get('reason', '')
        contributor.save(update_fields=['status', 'rejection_reason'])
        return DRFResponse({'status': 'rejected'})


# ============================================
# Public Contribute API (No Auth Required)
# ============================================

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def contribute_form(request, token):
    """
    Get the form for a contributor (public access via token).
    Returns the items they need to fill.
    """
    contributor = get_object_or_404(Contributor, invite_token=token)

    # Mark as accessed
    contributor.mark_accessed()

    # Get items for this contributor's entity
    items = contributor.entity.items.all().order_by('axis__order', 'order')

    # Get existing responses
    responses = contributor.responses.all()

    # Serialize items
    from apps.templates_app.serializers import ItemDetailSerializer
    items_data = ItemDetailSerializer(items, many=True).data
    responses_data = ResponseSerializer(responses, many=True).data

    # Get structure-derived data requirements (if structures exist)
    structure_hints = {}
    try:
        structures = ItemStructure.objects.filter(
            project=contributor.project,
            item__in=items
        ).select_related('item')
        for structure in structures:
            hints = []
            for comp in (structure.components or []):
                comp_type = comp.get('type', '')
                if comp_type == 'table':
                    hints.append({
                        'id': comp.get('id'),
                        'type': 'table',
                        'title': comp.get('title', ''),
                        'columns': comp.get('columns', []),
                        'suggested_input': 'table_dynamic',
                    })
                elif comp_type == 'chart':
                    hints.append({
                        'id': comp.get('id'),
                        'type': 'chart_data',
                        'title': comp.get('title', ''),
                        'chart_type': comp.get('chart_type', 'pie'),
                        'suggested_input': 'table_dynamic',
                    })
            if hints:
                structure_hints[structure.item_id] = {
                    'structure_id': str(structure.id),
                    'data_fields': hints,
                    'has_tables': any(h['type'] == 'table' for h in hints),
                    'has_charts': any(h['type'] == 'chart_data' for h in hints),
                }
    except Exception:
        pass

    return DRFResponse({
        'project': {
            'id': str(contributor.project.id),
            'name': contributor.project.name,
            'period': contributor.project.period,
            'deadline': contributor.project.deadline,
        },
        'organization': {
            'name': contributor.project.organization.name,
        },
        'entity': {
            'id': contributor.entity.id,
            'name': contributor.entity.name,
        },
        'contributor': {
            'id': str(contributor.id),
            'name': contributor.name,
            'status': contributor.status,
        },
        'items': items_data,
        'responses': responses_data,
        'progress': contributor.progress,
        'items_count': contributor.items_count,
        'completed_count': contributor.completed_items_count,
        'structure_hints': structure_hints,
    })


@api_view(['POST', 'PATCH'])
@permission_classes([permissions.AllowAny])
def contribute_save(request, token):
    """
    Save responses for a contributor (auto-save).
    Accepts a list of responses.
    """
    contributor = get_object_or_404(Contributor, invite_token=token)

    # Mark as accessed
    contributor.mark_accessed()

    responses_data = request.data.get('responses', [])

    saved_responses = []
    for response_data in responses_data:
        item_id = response_data.get('item_id') or response_data.get('item')
        value = response_data.get('value')
        attachments = response_data.get('attachments', [])

        if not item_id:
            continue

        # Verify item belongs to this entity
        try:
            item = contributor.entity.items.get(id=item_id)
        except Item.DoesNotExist:
            continue

        # Create or update response
        response, created = Response.objects.update_or_create(
            project=contributor.project,
            contributor=contributor,
            item=item,
            defaults={
                'value': {'value': value} if not isinstance(value, dict) else value,
                'attachments': attachments,
            }
        )
        saved_responses.append(response)

    # Update contributor status
    if contributor.progress == 100:
        contributor.status = 'submitted'
    else:
        contributor.status = 'in_progress'
    contributor.save(update_fields=['status'])

    return DRFResponse({
        'status': 'saved',
        'saved_count': len(saved_responses),
        'progress': contributor.progress,
    })


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def contribute_submit(request, token):
    """
    Submit the form (mark as complete).
    """
    contributor = get_object_or_404(Contributor, invite_token=token)

    contributor.status = 'submitted'
    contributor.submitted_at = timezone.now()
    contributor.save(update_fields=['status', 'submitted_at'])

    return DRFResponse({
        'status': 'submitted',
        'message': 'تم إرسال البيانات بنجاح'
    })


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def contribute_upload(request, token):
    """
    Upload an Excel file and parse it into Response objects.
    Supports .xlsx and .xls files.

    POST /api/reports/contribute/<token>/upload/
    FormData: file, item_id
    """
    contributor = get_object_or_404(Contributor, invite_token=token)

    if contributor.status in ('submitted', 'completed'):
        return DRFResponse(
            {'error': 'تم إرسال البيانات مسبقاً'},
            status=status.HTTP_400_BAD_REQUEST
        )

    item_id = request.data.get('item_id')
    file = request.FILES.get('file')

    if not file:
        return DRFResponse(
            {'error': 'لم يتم تحديد ملف'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if not item_id:
        return DRFResponse(
            {'error': 'لم يتم تحديد البند'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        item = Item.objects.get(id=item_id)
    except Item.DoesNotExist:
        return DRFResponse(
            {'error': 'البند غير موجود'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Parse Excel file
    if file.name.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active

            rows = []
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell) if cell else f'col_{i}' for i, cell in enumerate(row)]
                    continue
                if any(cell is not None for cell in row):
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        key = headers[col_idx] if col_idx < len(headers) else f'col_{col_idx}'
                        row_data[key] = cell
                    rows.append(row_data)

            wb.close()

            # Save as Response
            value_data = {'rows': rows, 'headers': headers, 'source': 'excel', 'filename': file.name}
            response_obj, created = Response.objects.update_or_create(
                project=contributor.project,
                contributor=contributor,
                item=item,
                defaults={
                    'value': value_data,
                    'attachments': [{'filename': file.name, 'size': file.size, 'type': 'excel'}]
                }
            )

            return DRFResponse({
                'status': 'uploaded',
                'filename': file.name,
                'size': file.size,
                'rows_count': len(rows),
                'headers': headers,
                'preview': rows[:5],
                'response_id': str(response_obj.id),
            })

        except Exception as e:
            return DRFResponse(
                {'error': f'فشل في قراءة الملف: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    else:
        # Non-Excel file: store as attachment reference
        response_obj, created = Response.objects.update_or_create(
            project=contributor.project,
            contributor=contributor,
            item=item,
            defaults={
                'value': {'value': file.name, 'source': 'file'},
                'attachments': [{'filename': file.name, 'size': file.size}]
            }
        )

        return DRFResponse({
            'status': 'uploaded',
            'filename': file.name,
            'size': file.size,
            'response_id': str(response_obj.id),
        })


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def contribute_excel_template(request, token, item_id):
    """
    تحميل قالب Excel فاضي بالأعمدة المحددة من Structure.

    GET /api/reports/contribute/<token>/excel-template/<item_id>/
    يرجع ملف .xlsx فاضي بأعمدة الجدول + الصفوف الثابتة إن وُجدت.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from apps.templates_app.models import TableDefinition

    contributor = get_object_or_404(Contributor, invite_token=token)
    item = get_object_or_404(Item, id=item_id)

    # Find table definitions for this item from Structure
    columns = []
    fixed_rows = []
    table_title = item.name

    # Source 1: ItemStructure components
    structure = ItemStructure.objects.filter(
        project=contributor.project, item=item
    ).first()
    if structure and structure.components:
        for comp in structure.components:
            if comp.get('type') == 'table':
                table_title = comp.get('title', item.name)
                # Get columns from component
                comp_columns = comp.get('columns', [])
                if comp_columns:
                    columns = comp_columns
                # Try to get from TableDefinition
                table_def_id = comp.get('table_def_id')
                if table_def_id and not columns:
                    try:
                        tdef = TableDefinition.objects.get(id=table_def_id)
                        columns = tdef.columns or []
                        fixed_rows = tdef.fixed_rows or []
                    except TableDefinition.DoesNotExist:
                        pass
                break  # Use first table found

    # Source 2: TableDefinition linked to item's axis
    if not columns:
        tdef = TableDefinition.objects.filter(
            template=contributor.project.template,
            axis=item.axis,
        ).first()
        if tdef:
            columns = tdef.columns or []
            fixed_rows = tdef.fixed_rows or []
            table_title = tdef.name

    # Fallback: basic template
    if not columns:
        columns = [
            {'name': 'العنصر', 'type': 'text'},
            {'name': 'القيمة', 'type': 'number'},
        ]

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'البيانات'
    ws.sheet_view.rightToLeft = True

    # Styles
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=table_title)
    title_cell.font = Font(name='Arial', bold=True, size=14, color='1A365D')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Instructions row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    instr_cell = ws.cell(row=2, column=1, value='يرجى تعبئة البيانات في الخلايا الفارغة أدناه ثم رفع الملف')
    instr_cell.font = Font(name='Arial', size=10, color='666666', italic=True)
    instr_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 25

    # Header row
    header_row = 3
    for col_idx, col_def in enumerate(columns, 1):
        col_name = col_def.get('name', col_def.get('key', f'عمود {col_idx}'))
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(15, len(col_name) * 2)
    ws.row_dimensions[header_row].height = 30

    # Fixed rows (for static tables)
    data_start_row = header_row + 1
    if fixed_rows:
        for row_idx, row_label in enumerate(fixed_rows, data_start_row):
            cell = ws.cell(row=row_idx, column=1, value=row_label)
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = cell_align
            cell.border = thin_border
            # Empty cells for other columns
            for col_idx in range(2, len(columns) + 1):
                empty_cell = ws.cell(row=row_idx, column=col_idx, value='')
                empty_cell.border = thin_border
                empty_cell.alignment = cell_align
    else:
        # Add 10 empty rows for dynamic tables
        for row_idx in range(data_start_row, data_start_row + 10):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value='')
                cell.border = thin_border
                cell.alignment = cell_align

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'template_{item.code}_{table_title[:30]}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ============================================
# Response ViewSet
# ============================================

class ResponseViewSet(viewsets.ModelViewSet):
    """API for Responses"""
    serializer_class = ResponseSerializer
    permission_classes = [permissions.AllowAny]  # For demo

    def get_queryset(self):
        queryset = Response.objects.all()

        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        contributor_id = self.request.query_params.get('contributor')
        if contributor_id:
            queryset = queryset.filter(contributor_id=contributor_id)

        item_id = self.request.query_params.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)

        return queryset


# ============================================
# Skeleton-First Workflow Views
# ============================================

class ItemStructureViewSet(viewsets.ModelViewSet):
    """
    API لهياكل البنود — Skeleton-First

    GET    /api/reports/structures/                    → قائمة كل الهياكل
    GET    /api/reports/structures/?project=UUID        → هياكل مشروع معين
    GET    /api/reports/structures/{id}/               → تفاصيل هيكل بند
    POST   /api/reports/structures/                    → إنشاء هيكل
    PATCH  /api/reports/structures/{id}/               → تعديل الهيكل
    """
    permission_classes = [permissions.AllowAny]
    serializer_class = ItemStructureSerializer

    def get_queryset(self):
        queryset = ItemStructure.objects.select_related(
            'project', 'item', 'item__axis'
        ).prefetch_related('generated_contents').all()

        # Filter by project
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # Filter by axis
        axis_id = self.request.query_params.get('axis')
        if axis_id:
            queryset = queryset.filter(item__axis_id=axis_id)

        # Filter by item
        item_id = self.request.query_params.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)

        return queryset

    def get_serializer_class(self):
        if self.action in ('update', 'partial_update'):
            return ItemStructureUpdateSerializer
        return ItemStructureSerializer

    @action(detail=False, methods=['post'])
    def init_from_template(self, request):
        """
        إنشاء هياكل البنود من القالب لمشروع معين
        """
        serializer = ItemStructureCreateFromTemplateSerializer(data=request.data)
        if not serializer.is_valid():
            return DRFResponse(serializer.errors, status=400)

        data = serializer.validated_data
        project = get_object_or_404(Project, id=data['project_id'])
        item_ids = data.get('item_ids', [])
        overwrite = data.get('overwrite', False)

        # Get items from template
        items_qs = Item.objects.filter(axis__template=project.template)
        if item_ids:
            items_qs = items_qs.filter(id__in=item_ids)

        created_count = 0
        skipped_count = 0
        errors = []

        for item in items_qs:
            try:
                # Check if already exists
                existing = ItemStructure.objects.filter(
                    project=project, item=item
                ).first()

                if existing and not overwrite:
                    skipped_count += 1
                    continue

                if existing and overwrite:
                    existing.delete()

                ItemStructure.create_from_template(project, item)
                created_count += 1

            except Exception as e:
                errors.append({
                    'item': item.code,
                    'error': str(e)
                })

        return DRFResponse({
            'status': 'success',
            'created': created_count,
            'skipped': skipped_count,
            'errors': errors,
            'message': f'تم إنشاء {created_count} هيكل، تم تخطي {skipped_count}',
        })

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """اعتماد الهيكل"""
        structure = self.get_object()
        structure.is_approved = True
        structure.save(update_fields=['is_approved', 'updated_at'])
        return DRFResponse({
            'status': 'success',
            'message': f'تم اعتماد هيكل البند {structure.item.code}'
        })

    @action(detail=True, methods=['get'])
    def context(self, request, pk=None):
        """
        الحصول على سياق فقرة معينة
        """
        structure = self.get_object()
        paragraph_id = request.query_params.get('paragraph', 'p1')
        ctx = structure.get_context_for_paragraph(paragraph_id)
        return DRFResponse(ctx)

    @action(detail=False, methods=['get'])
    def data_requirements(self, request):
        """
        تحليل هياكل البنود واستخراج متطلبات البيانات المطلوبة من المساهم
        """
        project_id = request.query_params.get('project')
        if not project_id:
            return DRFResponse({'error': 'project parameter مطلوب'}, status=400)

        structures = ItemStructure.objects.filter(
            project_id=project_id
        ).select_related('item', 'item__axis')

        requirements = []
        for structure in structures:
            item_reqs = {
                'item_id': structure.item_id,
                'item_code': structure.item.code,
                'item_name': structure.item.name,
                'axis_code': structure.item.axis.code,
                'fields': [],
            }

            has_tables = False
            has_charts = False
            has_paragraphs = False

            for comp in (structure.components or []):
                comp_type = comp.get('type', '')
                if comp_type == 'table':
                    has_tables = True
                    item_reqs['fields'].append({
                        'id': comp.get('id'),
                        'type': 'table',
                        'title': comp.get('title', 'جدول'),
                        'columns': comp.get('columns', []),
                        'data_source': comp.get('data_source', ''),
                        'input_method': 'table_dynamic',
                        'description': f'جدول: {comp.get("title", "")}',
                    })
                elif comp_type == 'chart':
                    has_charts = True
                    item_reqs['fields'].append({
                        'id': comp.get('id'),
                        'type': 'chart_data',
                        'title': comp.get('title', 'شكل'),
                        'chart_type': comp.get('chart_type', 'pie'),
                        'data_source': comp.get('data_source', ''),
                        'input_method': 'table_dynamic',
                        'description': f'بيانات شكل: {comp.get("title", "")}',
                    })
                elif comp_type == 'paragraph':
                    has_paragraphs = True

            # Always add basic numeric field for items with paragraphs
            if has_paragraphs:
                item_reqs['fields'].insert(0, {
                    'id': 'value',
                    'type': 'number',
                    'title': f'القيمة الحالية لـ {structure.item.name}',
                    'input_method': 'number',
                    'description': 'القيمة الرقمية الأساسية',
                })
                item_reqs['fields'].insert(1, {
                    'id': 'previous_value',
                    'type': 'number',
                    'title': 'القيمة السابقة (اختياري)',
                    'input_method': 'number',
                    'required': False,
                    'description': 'قيمة السنة السابقة للمقارنة',
                })

            item_reqs['summary'] = {
                'tables': has_tables,
                'charts': has_charts,
                'paragraphs': has_paragraphs,
                'total_fields': len(item_reqs['fields']),
            }

            if item_reqs['fields']:
                requirements.append(item_reqs)

        return DRFResponse({
            'project_id': project_id,
            'total_items': len(requirements),
            'requirements': requirements,
        })


class GeneratedContentViewSet(viewsets.ModelViewSet):
    """
    API للمحتويات المولّدة — فقرة بفقرة

    GET    /api/reports/generated-contents/                → قائمة المحتويات
    GET    /api/reports/generated-contents/?project=UUID   → محتويات مشروع
    GET    /api/reports/generated-contents/{id}/           → تفاصيل محتوى
    POST   /api/reports/generated-contents/{id}/edit/      → تعديل يدوي
    POST   /api/reports/generated-contents/{id}/approve/   → اعتماد
    POST   /api/reports/generated-contents/{id}/regenerate/ → إعادة توليد
    """
    permission_classes = [permissions.AllowAny]
    serializer_class = GeneratedContentSerializer

    def get_queryset(self):
        queryset = GeneratedContent.objects.select_related(
            'item_structure', 'item_structure__item',
            'item_structure__item__axis', 'project'
        ).all()

        # Filter by project
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # Filter by structure
        structure_id = self.request.query_params.get('structure')
        if structure_id:
            queryset = queryset.filter(item_structure_id=structure_id)

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        # Filter by item
        item_id = self.request.query_params.get('item')
        if item_id:
            queryset = queryset.filter(item_structure__item_id=item_id)

        return queryset

    @action(detail=True, methods=['post'])
    def edit(self, request, pk=None):
        """
        تعديل يدوي لمحتوى فقرة
        """
        content_obj = self.get_object()
        serializer = GeneratedContentEditSerializer(data=request.data)
        if not serializer.is_valid():
            return DRFResponse(serializer.errors, status=400)

        content_obj.edit(
            serializer.validated_data['content'],
            user=request.user if request.user.is_authenticated else None
        )

        return DRFResponse(GeneratedContentSerializer(content_obj).data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """اعتماد محتوى فقرة"""
        content_obj = self.get_object()
        content_obj.approve(
            user=request.user if request.user.is_authenticated else None
        )
        return DRFResponse({
            'status': 'success',
            'message': f'تم اعتماد {content_obj.component_id}'
        })

    @action(detail=True, methods=['post'])
    def regenerate(self, request, pk=None):
        """
        إعادة توليد فقرة واحدة
        """
        content_obj = self.get_object()
        model = request.data.get('model', 'cli')
        extra_instructions = request.data.get('extra_instructions', '')
        user = request.user if request.user.is_authenticated else None

        content_obj.regenerate()

        from .text_generator import TextGenerator

        generator = TextGenerator(content_obj.project, model=model)
        result = generator.generate_paragraph(
            content_obj,
            extra_instructions=extra_instructions,
            user=user,
        )

        if result.get('success'):
            content_obj.refresh_from_db()
            return DRFResponse(GeneratedContentSerializer(content_obj).data)
        else:
            return DRFResponse({
                'status': 'failed',
                'error': result.get('error', 'فشل التوليد'),
                'content_id': str(content_obj.id),
            }, status=500)


class TableDataViewSet(viewsets.ModelViewSet):
    """
    API لبيانات الجداول

    GET    /api/reports/table-data/                → قائمة البيانات
    GET    /api/reports/table-data/?project=UUID   → بيانات مشروع
    PATCH  /api/reports/table-data/{id}/           → تعديل بيانات
    POST   /api/reports/table-data/{id}/update_rows/ → تعديل صفوف
    """
    permission_classes = [permissions.AllowAny]
    serializer_class = TableDataSerializer

    def get_queryset(self):
        queryset = TableData.objects.select_related(
            'project', 'contributor', 'table_definition'
        ).all()

        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        item_id = self.request.query_params.get('item')
        if item_id:
            queryset = queryset.filter(
                table_definition__axis__items__id=item_id
            ).distinct()

        table_def_id = self.request.query_params.get('table_definition')
        if table_def_id:
            queryset = queryset.filter(table_definition_id=table_def_id)

        return queryset

    @action(detail=True, methods=['post'])
    def update_rows(self, request, pk=None):
        """
        تعديل صفوف جدول (تعديل خلايا / إضافة / حذف)
        """
        obj = self.get_object()
        rows = request.data.get('rows')
        if rows is None:
            return DRFResponse({'error': 'حقل rows مطلوب'}, status=400)
        obj.rows = rows
        obj.save(update_fields=['rows', 'updated_at'])
        return DRFResponse(TableDataSerializer(obj).data)


# ============================================
# Skeleton & Text Generation API Endpoints
# ============================================

@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def build_skeleton(request):
    """
    بناء الهيكل (Skeleton) لمشروع

    POST /api/reports/build-skeleton/
    {
        "project_id": "uuid",
        "item_ids": [1, 2, 3]  // اختياري
    }
    """
    serializer = SkeletonBuildRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return DRFResponse(serializer.errors, status=400)

    data = serializer.validated_data
    project = get_object_or_404(Project, id=data['project_id'])
    item_ids = data.get('item_ids', [])

    # Get items
    items_qs = Item.objects.filter(axis__template=project.template).order_by('axis__order', 'order')
    if item_ids:
        items_qs = items_qs.filter(id__in=item_ids)

    structures_created = 0
    contents_created = 0
    errors = []

    for item in items_qs:
        try:
            # 1. Create or get ItemStructure
            structure, created = ItemStructure.objects.get_or_create(
                project=project,
                item=item,
                defaults={
                    'components': [],
                    'source': 'template',
                }
            )

            if created:
                # Populate from template
                structure = ItemStructure.create_from_template(project, item)
                structures_created += 1
            elif not structure.components:
                # Repopulate if empty
                temp = ItemStructure.create_from_template(project, item)
                structure.components = temp.components
                structure.save()
                # Delete the temp duplicate
                ItemStructure.objects.filter(
                    project=project, item=item
                ).exclude(id=structure.id).delete()

            # 2. Create GeneratedContent for each paragraph
            for comp in structure.get_paragraphs():
                gc, gc_created = GeneratedContent.objects.get_or_create(
                    item_structure=structure,
                    component_id=comp['id'],
                    defaults={
                        'project': project,
                        'status': 'not_started',
                    }
                )
                if gc_created:
                    contents_created += 1

        except Exception as e:
            errors.append({
                'item': item.code,
                'error': str(e)
            })

    return DRFResponse({
        'status': 'success',
        'structures_created': structures_created,
        'contents_created': contents_created,
        'total_items': items_qs.count(),
        'errors': errors,
        'message': f'تم بناء الهيكل: {structures_created} هيكل، {contents_created} فقرة جاهزة للتوليد',
    })


@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def generate_text(request):
    """
    توليد نصوص AI لفقرات الهيكل

    POST /api/reports/generate-text/
    {
        "project_id": "uuid",
        "item_id": 5,           // اختياري — بند معين
        "component_id": "p1",   // اختياري — فقرة معينة
        "structure_id": "uuid", // اختياري — هيكل معين
        "model": "cli",
        "extra_instructions": ""
    }
    """
    serializer = TextGenerateRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return DRFResponse(serializer.errors, status=400)

    data = serializer.validated_data
    project = get_object_or_404(Project, id=data['project_id'])
    model = data.get('model', 'cli')
    extra_instructions = data.get('extra_instructions', '')
    user = request.user if request.user.is_authenticated else None

    # Find target GeneratedContent records
    queryset = GeneratedContent.objects.filter(project=project)

    if data.get('structure_id'):
        queryset = queryset.filter(item_structure_id=data['structure_id'])
    elif data.get('item_id'):
        queryset = queryset.filter(item_structure__item_id=data['item_id'])

    if data.get('component_id'):
        queryset = queryset.filter(component_id=data['component_id'])

    # Filter eligible records
    targets = queryset.filter(status__in=['not_started', 'failed'])
    count = targets.count()

    if count == 0:
        return DRFResponse({
            'status': 'info',
            'message': 'لا توجد فقرات جاهزة للتوليد. تأكد من بناء الهيكل أولاً.',
        })

    # Mark all as generating
    targets.update(status='generating')

    # Call TextGenerator
    from .text_generator import TextGenerator

    generator = TextGenerator(project, model=model)
    results = {
        'generated': [],
        'failed': [],
        'total_cost': 0,
        'total_duration_ms': 0,
    }

    for gc in targets.select_related(
        'item_structure', 'item_structure__item', 'item_structure__item__axis'
    ):
        result = generator.generate_paragraph(
            gc,
            extra_instructions=extra_instructions,
            user=user,
        )
        if result.get('success'):
            results['generated'].append({
                'id': str(gc.id),
                'component_id': gc.component_id,
                'item_code': gc.item_structure.item.code,
                'status': 'generated',
            })
            results['total_cost'] += result.get('cost', 0)
            results['total_duration_ms'] += result.get('duration_ms', 0)
        else:
            results['failed'].append({
                'id': str(gc.id),
                'component_id': gc.component_id,
                'item_code': gc.item_structure.item.code,
                'error': result.get('error', 'Unknown'),
            })

    gen_count = len(results['generated'])
    fail_count = len(results['failed'])

    return DRFResponse({
        'status': 'completed' if not results['failed'] else 'partial',
        'generated_count': gen_count,
        'failed_count': fail_count,
        'total_cost': results['total_cost'],
        'total_duration_ms': results['total_duration_ms'],
        'generated': results['generated'],
        'failed': results['failed'],
        'message': f'تم توليد {gen_count} فقرة من أصل {count}',
    })


def _build_preview_html(project, structures):
    """Generate a simple HTML preview of the project skeleton."""
    items_html = ''
    for struct in structures.select_related('item', 'item__axis').order_by('item__axis__order', 'item__order'):
        item = struct.item
        components = struct.components or []
        comps_html = ''
        for comp in components:
            ctype = comp.get('type', 'paragraph')
            title = comp.get('title') or comp.get('description') or ''
            if ctype == 'paragraph':
                comps_html += f'''
                <div style="margin:10px 0;padding:12px 16px;background:#f5f3ff;border-right:4px solid #7c3aed;border-radius:6px;">
                  <div style="font-size:11px;color:#7c3aed;font-weight:600;margin-bottom:6px;">📝 فقرة نصية — ينتظر الذكاء الاصطناعي</div>
                  <div style="font-size:12px;color:#6d28d9;">{title}</div>
                  <div style="margin-top:8px;display:flex;gap:4px;flex-wrap:wrap;">
                    {''.join('<span style="display:inline-block;height:10px;background:#ddd8fe;border-radius:4px;margin:2px 0;" style="width:{w}%"></span>'.replace('{w}', str(w)) for w in [90, 75, 85, 60]) }
                  </div>
                </div>'''
            elif ctype == 'table':
                cols = comp.get('columns') or []
                cols_header = ''.join(f'<th style="padding:8px 12px;background:#f0fdf4;border:1px solid #bbf7d0;font-size:12px;">{c.get("name","") if isinstance(c,dict) else c}</th>' for c in (cols or ['العمود 1', 'العمود 2', 'العمود 3']))
                has_data = comp.get('has_data', False)
                status_badge = '<span style="font-size:10px;background:#dcfce7;color:#166534;padding:2px 8px;border-radius:12px;">✓ بيانات جاهزة</span>' if has_data else '<span style="font-size:10px;background:#fef9c3;color:#854d0e;padding:2px 8px;border-radius:12px;">⚠ بيانات ناقصة</span>'
                comps_html += f'''
                <div style="margin:10px 0;">
                  <div style="font-size:11px;color:#15803d;font-weight:600;margin-bottom:6px;">📋 جدول بيانات {status_badge}</div>
                  <div style="font-size:12px;color:#374151;margin-bottom:8px;">{title}</div>
                  <table style="width:100%;border-collapse:collapse;font-size:12px;">
                    <thead><tr>{cols_header if cols else '<th style="padding:8px 12px;background:#f0fdf4;border:1px solid #bbf7d0;">البيانات</th>'}</tr></thead>
                    <tbody>
                      <tr>{''.join('<td style="padding:8px 12px;border:1px solid #e5e7eb;color:#9ca3af;font-style:italic;">—</td>' for _ in (cols or ['']))}</tr>
                    </tbody>
                  </table>
                </div>'''
            elif ctype == 'chart':
                chart_type = comp.get('chart_type', 'bar')
                chart_labels = {'pie': 'دائري', 'bar': 'أعمدة', 'line': 'خطي', 'area': 'مساحي'}
                comps_html += f'''
                <div style="margin:10px 0;padding:12px 16px;background:#fffbeb;border:1px dashed #fbbf24;border-radius:6px;text-align:center;">
                  <div style="font-size:11px;color:#92400e;font-weight:600;margin-bottom:6px;">📊 شكل بياني — {chart_labels.get(chart_type, chart_type)}</div>
                  <div style="font-size:12px;color:#78350f;">{title}</div>
                  <div style="margin-top:12px;padding:20px;background:#fef3c7;border-radius:4px;color:#92400e;font-size:13px;">[ رسم بياني ]</div>
                </div>'''

        items_html += f'''
        <div style="margin-bottom:24px;background:#fff;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;">
          <div style="padding:14px 20px;background:#f9fafb;border-bottom:1px solid #e5e7eb;">
            <div style="font-weight:700;color:#111827;font-size:14px;">{item.code}. {item.name}</div>
            <div style="font-size:12px;color:#6b7280;margin-top:2px;">{item.axis.name}</div>
          </div>
          <div style="padding:16px 20px;">{comps_html or '<p style="color:#9ca3af;font-size:13px;">لا توجد مكونات</p>'}</div>
        </div>'''

    if not items_html:
        return '<div dir="rtl" style="text-align:center;padding:60px 20px;color:#9ca3af;font-family:Cairo,sans-serif;">لم يُبنَ الهيكل بعد — اضغط «بناء الهيكل» أولاً</div>'

    return f'''<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head>
<meta charset="UTF-8">
<style>
  body {{ font-family: "Cairo", "Segoe UI", Arial, sans-serif; background: #f9fafb; color: #111827; margin: 0; padding: 24px; }}
  h1 {{ font-size: 20px; font-weight: 700; color: #1e40af; margin-bottom: 4px; }}
  .subtitle {{ font-size: 13px; color: #6b7280; margin-bottom: 24px; }}
</style>
</head>
<body>
  <h1>{project.name}</h1>
  <div class="subtitle">هيكل التقرير — الفترة: {project.period or ''}</div>
  {items_html}
</body>
</html>'''


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def project_skeleton_status(request, project_id):
    """
    حالة الهيكل لمشروع معين

    GET /api/reports/projects/{project_id}/skeleton-status/
    """
    project = get_object_or_404(Project, id=project_id)

    structures = ItemStructure.objects.filter(project=project)
    contents = GeneratedContent.objects.filter(project=project)

    total_items = Item.objects.filter(axis__template=project.template).count()

    # Build preview HTML
    preview_html = _build_preview_html(project, structures)

    return DRFResponse({
        'project_id': str(project.id),
        'project_name': project.name,
        'total_items': total_items,
        'structures_count': structures.count(),
        'structures_approved': structures.filter(is_approved=True).count(),
        'contents_total': contents.count(),
        'contents_by_status': {
            'not_started': contents.filter(status='not_started').count(),
            'generating': contents.filter(status='generating').count(),
            'generated': contents.filter(status='generated').count(),
            'edited': contents.filter(status='edited').count(),
            'approved': contents.filter(status='approved').count(),
            'failed': contents.filter(status='failed').count(),
        },
        'progress': {
            'structure': int(structures.count() / total_items * 100) if total_items > 0 else 0,
            'generation': int(contents.filter(
                status__in=['generated', 'edited', 'approved']
            ).count() / contents.count() * 100) if contents.count() > 0 else 0,
            'approval': int(contents.filter(
                status='approved'
            ).count() / contents.count() * 100) if contents.count() > 0 else 0,
        },
        'preview_html': preview_html,
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def analyze_previous_report(request):
    """
    تحليل تقرير سابق (Word) واستخراج هيكله

    POST /api/reports/analyze-report/
    Body: multipart/form-data with 'file' field
    """
    from .report_analyzer import analyze_uploaded_report

    if 'file' not in request.FILES:
        return DRFResponse(
            {'error': 'يرجى رفع ملف Word (.docx)'},
            status=status.HTTP_400_BAD_REQUEST
        )

    uploaded_file = request.FILES['file']
    if not uploaded_file.name.endswith('.docx'):
        return DRFResponse(
            {'error': 'الملف يجب أن يكون بصيغة .docx'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        file_bytes = uploaded_file.read()
        result = analyze_uploaded_report(file_bytes)

        if 'error' in result and result['error']:
            return DRFResponse(
                {'error': f'فشل في تحليل الملف: {result["error"]}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return DRFResponse(result)

    except Exception as e:
        return DRFResponse(
            {'error': f'حدث خطأ: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
