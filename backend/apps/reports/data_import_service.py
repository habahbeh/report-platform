"""
Data Import Service
خدمة استيراد البيانات من ملفات Excel وربطها بالبنود

المميزات:
- قراءة ملفات Excel تلقائياً
- ربط البيانات بالبنود المناسبة
- حساب المؤشرات والنسب
- تحديث ItemDraft بالقيم
"""

import os
import re
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter

from django.db import transaction

from apps.data_collection.models import DataCollectionPeriod
from apps.templates_app.models import Item, Axis
from apps.reports.models import ItemDraft


# ============================================
# Mapping: Item Code → Data Source
# ============================================

# هذا الـ mapping يربط كل بند بمصدر بياناته
ITEM_DATA_MAPPING = {
    # المحور 1: الاعتمادية وضمان الجودة
    '1.5': {
        'source': 'manual',  # يحتاج إدخال يدوي (تصنيفات عالمية)
        'description': 'تصنيف QS - يُدخل يدوياً من موقع QS',
    },
    '1.6': {
        'source': 'manual',
        'description': 'تصنيف THE - يُدخل يدوياً',
    },
    '1.7': {
        'source': 'manual',
        'description': 'تصنيف Shanghai - يُدخل يدوياً',
    },
    '1.9': {
        'source': 'calculated',
        'formula': '(non_jordanian_students / total_students) * 100',
        'depends_on': ['1.8'],  # يعتمد على بند 1.8
    },
    
    # المحور 2: التدريس
    '2.2': {
        'source': 'calculated',
        'description': 'إجمالي الزيادة - يُحسب من بيانات الهيئة التدريسية',
    },
    
    # المحور 3: البحث العلمي
    '3.4': {
        'source': 'file',
        'folder': '04_عمادة_البحث_العلمي',
        'file': 'الابحاث_المنشورة.xlsx',
        'value_cell': 'calculated',  # نسبة
    },
    '3.5': {
        'source': 'file',
        'folder': '04_عمادة_البحث_العلمي',
        'file': 'الابحاث_المنشورة.xlsx',
        'value_cell': 'calculated',
    },
    
    # المحور 4: إدارة الموارد البشرية والمالية
    '4.1': {
        'source': 'file',
        'folder': '14_الدائرة_المالية',
        'file': 'الموازنة.xlsx',
        'value_type': 'percentage',
        'description': 'نسبة الانحرافات - تُحسب من الموازنة',
    },
    '4.2': {
        'source': 'file',
        'folder': '14_الدائرة_المالية',
        'files': ['الايرادات.xlsx', 'المصروفات.xlsx'],
        'value_type': 'percentage',
        'description': 'نسبة تغطية الإيرادات للنفقات',
    },
    '4.3': {
        'source': 'file',
        'folder': '25_مكتب_الدراسات_والاستشارات',
        'description': 'إيرادات الربط مع الصناعة',
    },
    '4.4': {
        'source': 'manual',
        'description': 'صافي الدخل من الوحدات الإنتاجية',
    },
    '4.5': {
        'source': 'manual',
        'description': 'صافي الدخل من الأنشطة الاستثمارية',
    },
    '4.6': {
        'source': 'file',
        'folder': '14_الدائرة_المالية',
        'files': ['الايرادات.xlsx', 'المصروفات.xlsx'],
        'description': 'الوفر/العجز النقدي السنوي',
    },
    '4.7': {
        'source': 'manual',
        'description': 'العجز المتراكم',
    },
    '4.8': {
        'source': 'manual',
        'description': 'مساهمة صندوق نهاية الخدمة',
    },
    '4.9': {
        'source': 'manual',
        'description': 'مساهمة صندوق التأمين الصحي',
    },
    '4.10': {
        'source': 'file',
        'folder': '05_شؤون_الموظفين',
        'description': 'نسبة المكافآت للإداريين',
    },
    '4.11': {
        'source': 'file',
        'folder': '09_مركز_التطوير_الأكاديمي',
        'description': 'عدد الدورات للإداريين',
    },
    '4.12': {
        'source': 'calculated',
        'description': 'نسبة الإداريين لأعضاء هيئة التدريس',
    },
    '4.13': {
        'source': 'manual',
        'description': 'القروض غير المسددة',
    },
    '4.14': {
        'source': 'manual',
        'description': 'الفوائد البنكية',
    },
    
    # المحور 6: خدمة المجتمع
    '6.2': {
        'source': 'calculated',
        'depends_on': ['6.1'],
        'description': 'قيمة المشاريع - تُحسب من تفاصيل 6.1',
    },
    '6.4': {
        'source': 'calculated',
        'depends_on': ['6.3'],
        'description': 'عدد الطلبة المشاركين',
    },
}


class DataImportService:
    """خدمة استيراد البيانات"""
    
    def __init__(self, period: DataCollectionPeriod, data_folder: str):
        self.period = period
        self.data_folder = Path(data_folder)
        self.results = {
            'imported': [],
            'skipped': [],
            'errors': [],
            'manual_needed': [],
        }
    
    def import_all(self) -> Dict[str, Any]:
        """استيراد كل البيانات المتاحة"""
        
        # 1. استيراد البيانات المالية (المحور 4)
        self._import_financial_data()
        
        # 2. استيراد باقي البيانات من الملفات
        self._import_from_files()
        
        # 3. حساب البنود المحسوبة
        self._calculate_derived_items()
        
        # 4. تحديد البنود اليدوية
        self._identify_manual_items()
        
        return self.results
    
    def _import_financial_data(self):
        """استيراد بيانات الدائرة المالية"""
        financial_folder = self.data_folder / '14_الدائرة_المالية'
        
        if not financial_folder.exists():
            self.results['errors'].append({
                'item': 'المحور 4',
                'error': 'مجلد الدائرة المالية غير موجود'
            })
            return
        
        try:
            # 4.1: نسبة الانحرافات من الموازنة
            budget_file = financial_folder / 'الموازنة.xlsx'
            if budget_file.exists():
                deviation = self._calculate_budget_deviation(budget_file)
                self._update_item_draft('4.1', deviation, source_file=str(budget_file))
            
            # 4.2: نسبة تغطية الإيرادات للنفقات
            revenue_file = financial_folder / 'الايرادات.xlsx'
            expense_file = financial_folder / 'المصروفات.xlsx'
            if revenue_file.exists() and expense_file.exists():
                coverage = self._calculate_coverage_ratio(revenue_file, expense_file)
                self._update_item_draft('4.2', coverage, 
                    source_file=f"{revenue_file}, {expense_file}",
                    table_data=self._get_coverage_table(revenue_file, expense_file))
            
            # 4.6: الوفر/العجز النقدي
            if revenue_file.exists() and expense_file.exists():
                surplus = self._calculate_surplus(revenue_file, expense_file)
                self._update_item_draft('4.6', surplus)
                
        except Exception as e:
            self.results['errors'].append({
                'item': 'المحور 4 المالي',
                'error': str(e)
            })
    
    def _calculate_budget_deviation(self, budget_file: Path) -> float:
        """حساب نسبة الانحراف من ملف الموازنة"""
        wb = openpyxl.load_workbook(budget_file)
        ws = wb.active
        
        total_estimated = 0
        total_actual = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:  # المقدر والفعلي
                try:
                    estimated = float(row[1]) if row[1] else 0
                    actual = float(row[2]) if row[2] else 0
                    total_estimated += estimated
                    total_actual += actual
                except (ValueError, TypeError):
                    continue
        
        if total_estimated > 0:
            deviation = ((total_actual - total_estimated) / total_estimated) * 100
            return round(deviation, 2)
        return 0
    
    def _calculate_coverage_ratio(self, revenue_file: Path, expense_file: Path) -> float:
        """حساب نسبة تغطية الإيرادات للنفقات"""
        total_revenue = self._sum_amounts(revenue_file)
        total_expense = self._sum_amounts(expense_file)
        
        if total_expense > 0:
            return round((total_revenue / total_expense) * 100, 2)
        return 0
    
    def _calculate_surplus(self, revenue_file: Path, expense_file: Path) -> float:
        """حساب الوفر/العجز"""
        total_revenue = self._sum_amounts(revenue_file)
        total_expense = self._sum_amounts(expense_file)
        return total_revenue - total_expense
    
    def _sum_amounts(self, file_path: Path) -> float:
        """جمع المبالغ من ملف Excel"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # العمود الثاني (المبلغ)
                try:
                    total += float(row[1])
                except (ValueError, TypeError):
                    continue
        return total
    
    def _get_coverage_table(self, revenue_file: Path, expense_file: Path) -> List[Dict]:
        """إنشاء جدول المقارنة"""
        revenues = []
        wb = openpyxl.load_workbook(revenue_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                revenues.append({'البند': row[0], 'المبلغ': row[1]})
        
        return revenues
    
    def _import_from_files(self):
        """استيراد البيانات من الملفات المحددة في الـ mapping"""
        for item_code, config in ITEM_DATA_MAPPING.items():
            if config.get('source') != 'file':
                continue
            
            folder = config.get('folder')
            if not folder:
                continue
            
            folder_path = self.data_folder / folder
            if not folder_path.exists():
                self.results['skipped'].append({
                    'item': item_code,
                    'reason': f'المجلد غير موجود: {folder}'
                })
                continue
            
            # محاولة استيراد البيانات
            try:
                file_name = config.get('file')
                if file_name:
                    file_path = folder_path / file_name
                    if file_path.exists():
                        data = self._extract_data_from_file(file_path, config)
                        if data:
                            self._update_item_draft(item_code, data.get('value'), 
                                source_file=str(file_path),
                                table_data=data.get('table'))
            except Exception as e:
                self.results['errors'].append({
                    'item': item_code,
                    'error': str(e)
                })
    
    def _extract_data_from_file(self, file_path: Path, config: Dict) -> Optional[Dict]:
        """استخراج البيانات من ملف Excel"""
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # قراءة كل البيانات
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        
        headers = rows[0]
        data_rows = rows[1:]
        
        # حساب القيمة حسب النوع
        value = None
        if config.get('value_type') == 'count':
            value = len(data_rows)
        elif config.get('value_type') == 'sum':
            col_index = config.get('sum_column', 1)
            value = sum(float(row[col_index]) for row in data_rows if row[col_index])
        else:
            value = len(data_rows)  # افتراضي: عدد الصفوف
        
        # إنشاء الجدول
        table = []
        for row in data_rows:
            row_dict = {}
            for i, header in enumerate(headers):
                if header and i < len(row):
                    row_dict[str(header)] = row[i]
            if row_dict:
                table.append(row_dict)
        
        return {
            'value': value,
            'table': table
        }
    
    def _calculate_derived_items(self):
        """حساب البنود المشتقة"""
        for item_code, config in ITEM_DATA_MAPPING.items():
            if config.get('source') != 'calculated':
                continue
            
            depends_on = config.get('depends_on', [])
            
            # التحقق من توفر البيانات المطلوبة
            all_available = True
            for dep_code in depends_on:
                draft = self._get_item_draft(dep_code)
                if not draft or draft.current_value is None:
                    all_available = False
                    break
            
            if not all_available:
                self.results['skipped'].append({
                    'item': item_code,
                    'reason': f'يعتمد على بنود غير متوفرة: {depends_on}'
                })
                continue
            
            # حساب القيمة (يحتاج تخصيص لكل بند)
            # هذا placeholder - يحتاج implementation محدد
    
    def _identify_manual_items(self):
        """تحديد البنود التي تحتاج إدخال يدوي"""
        for item_code, config in ITEM_DATA_MAPPING.items():
            if config.get('source') == 'manual':
                draft = self._get_item_draft(item_code)
                if not draft or draft.current_value is None:
                    self.results['manual_needed'].append({
                        'item': item_code,
                        'description': config.get('description', ''),
                    })
    
    def _get_item_draft(self, item_code: str) -> Optional[ItemDraft]:
        """الحصول على مسودة البند"""
        try:
            item = Item.objects.get(code=item_code, axis__template=self.period.template)
            draft, _ = ItemDraft.objects.get_or_create(period=self.period, item=item)
            return draft
        except Item.DoesNotExist:
            return None
    
    @transaction.atomic
    def _update_item_draft(self, item_code: str, value: Any, 
                          source_file: str = None, table_data: List = None):
        """تحديث مسودة البند بالبيانات"""
        draft = self._get_item_draft(item_code)
        if not draft:
            self.results['errors'].append({
                'item': item_code,
                'error': 'البند غير موجود في القالب'
            })
            return
        
        # تحديث القيمة
        if value is not None:
            draft.current_value = value
        
        # تحديث الجدول
        if table_data:
            draft.table_data = table_data
        
        draft.save()
        
        self.results['imported'].append({
            'item': item_code,
            'value': value,
            'has_table': bool(table_data),
        })


def import_all_data(period_id: int, data_folder: str) -> Dict[str, Any]:
    """
    دالة مساعدة لاستيراد كل البيانات
    
    Args:
        period_id: معرف فترة الجمع
        data_folder: مسار مجلد البيانات
    
    Returns:
        نتائج الاستيراد
    """
    period = DataCollectionPeriod.objects.get(id=period_id)
    service = DataImportService(period, data_folder)
    return service.import_all()
