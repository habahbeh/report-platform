#!/usr/bin/env python3
"""
Complete all data linking:
1. Link ALL tables to correct ItemComponents
2. Link ALL charts to correct ItemComponents
3. Generate detailed progress report
"""

import os
import sys
import json
import re
import django

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from apps.templates_app.models import Template, Item, ItemComponent, Axis

# Paths
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
EXTRACTED_DIR = os.path.join(DATA_DIR, 'extracted')


def load_extracted_data():
    """Load all extracted data."""
    with open(os.path.join(EXTRACTED_DIR, 'all_tables.json'), 'r', encoding='utf-8') as f:
        tables = json.load(f)
    
    with open(os.path.join(EXTRACTED_DIR, 'all_charts.json'), 'r', encoding='utf-8') as f:
        charts = json.load(f)
    
    return tables, charts


def link_tables(tables, template):
    """Link all tables to ItemComponents."""
    print("\n📊 ربط الجداول...")
    
    # Group tables by item_code
    tables_by_item = {}
    for table in tables:
        item_code = table.get('item_code')
        if item_code:
            if item_code not in tables_by_item:
                tables_by_item[item_code] = []
            tables_by_item[item_code].append(table)
    
    # Also create index by table_code
    tables_by_code = {}
    for table in tables:
        code = table.get('table_code')
        if code:
            tables_by_code[code] = table
    
    linked = 0
    already_linked = 0
    not_found = 0
    
    for item in Item.objects.filter(axis__template=template):
        item_tables = tables_by_item.get(item.code, [])
        table_idx = 0
        
        for comp in item.components.filter(component_type='table').order_by('order'):
            config = comp.config or {}
            
            # Skip if already has data
            if config.get('extracted_data') and config['extracted_data'].get('data'):
                already_linked += 1
                continue
            
            # Try to find by table_code first
            table_code = config.get('table_code')
            if table_code and table_code in tables_by_code:
                table = tables_by_code[table_code]
                config['extracted_data'] = {
                    'headers': table['headers'],
                    'data': table['data'][:100],
                    'total_rows': len(table['data']),
                    'source': 'by_code'
                }
                comp.config = config
                comp.save()
                linked += 1
                continue
            
            # Try to find by item index
            if table_idx < len(item_tables):
                table = item_tables[table_idx]
                config['extracted_data'] = {
                    'headers': table['headers'],
                    'data': table['data'][:100],
                    'total_rows': len(table['data']),
                    'source': 'by_index'
                }
                comp.config = config
                comp.save()
                linked += 1
                table_idx += 1
                continue
            
            not_found += 1
    
    print(f"   ✓ مرتبط جديد: {linked}")
    print(f"   ✓ مرتبط سابقاً: {already_linked}")
    print(f"   ⚠ غير موجود: {not_found}")
    
    return linked + already_linked


def link_charts(charts, template):
    """Link all charts to ItemComponents based on content analysis."""
    print("\n📈 ربط الرسوم البيانية...")
    
    # Analyze chart content to map to items
    chart_to_item = []
    
    for chart in charts:
        series = chart.get('series', [])
        if not series:
            continue
        
        cats = str(series[0].get('categories', []))
        vals = series[0].get('values', [])
        title = chart.get('title') or ''
        
        item_code = None
        
        # Detect item based on content
        if 'طلبة' in cats or 'طلب' in title.lower():
            if 'أجانب' in cats or 'أردن' in cats:
                item_code = '1.9'  # Student ratios
        elif 'كلية' in cats:
            if any(v > 500 for v in vals if isinstance(v, (int, float))):
                item_code = '2.2'  # Faculty distribution
            else:
                item_code = '1.9'
        elif any(str(y).startswith('20') for y in series[0].get('categories', [])):
            if 'بحث' in title or 'أبحاث' in title:
                item_code = '3.1'
            elif 'وثائق' in title or 'تعاون' in title:
                item_code = '1.10'
            else:
                item_code = '1.9'
        elif 'أستاذ' in cats:
            item_code = '2.2'
        elif 'ماجستير' in cats or 'بكالوريوس' in cats:
            item_code = '3.4'
        elif 'دعم' in title or 'مشاريع' in cats:
            item_code = '3.7'
        elif 'ورقة بحثية' in cats or 'مؤتمر' in cats:
            item_code = '3.8'
        elif 'جائزة' in cats:
            item_code = '3.8'
        
        chart_to_item.append({
            'chart': chart,
            'item_code': item_code
        })
    
    # Group by item
    charts_by_item = {}
    for mapping in chart_to_item:
        item_code = mapping['item_code']
        if item_code:
            if item_code not in charts_by_item:
                charts_by_item[item_code] = []
            charts_by_item[item_code].append(mapping['chart'])
    
    # Link to components
    linked = 0
    already_linked = 0
    
    for item in Item.objects.filter(axis__template=template):
        item_charts = charts_by_item.get(item.code, [])
        chart_idx = 0
        
        for comp in item.components.filter(component_type='chart').order_by('order'):
            config = comp.config or {}
            
            if config.get('chart_data') and config['chart_data'].get('series'):
                already_linked += 1
                continue
            
            if chart_idx < len(item_charts):
                chart = item_charts[chart_idx]
                config['chart_data'] = {
                    'type': chart.get('type'),
                    'title': chart.get('title'),
                    'series': chart.get('series', [])
                }
                comp.config = config
                comp.save()
                linked += 1
                chart_idx += 1
    
    print(f"   ✓ مرتبط جديد: {linked}")
    print(f"   ✓ مرتبط سابقاً: {already_linked}")
    
    return linked + already_linked


def generate_checklist(template):
    """Generate detailed progress checklist."""
    print("\n📋 إنشاء قائمة التقدم التفصيلية...")
    
    checklist = {
        'summary': {
            'total_items': 0,
            'total_components': 0,
            'texts': {'total': 0, 'with_content': 0},
            'tables': {'total': 0, 'with_data': 0},
            'charts': {'total': 0, 'with_data': 0},
        },
        'axes': []
    }
    
    for axis in Axis.objects.filter(template=template).order_by('order'):
        axis_data = {
            'code': axis.code,
            'name': axis.name,
            'items': [],
            'totals': {
                'texts': 0, 'texts_ok': 0,
                'tables': 0, 'tables_ok': 0,
                'charts': 0, 'charts_ok': 0,
            }
        }
        
        for item in Item.objects.filter(axis=axis).order_by('order'):
            item_data = {
                'code': item.code,
                'name': item.name[:40],
                'components': [],
                'texts': 0, 'texts_ok': 0,
                'tables': 0, 'tables_ok': 0,
                'charts': 0, 'charts_ok': 0,
            }
            
            for comp in item.components.all().order_by('order'):
                config = comp.config or {}
                has_data = False
                
                if comp.component_type == 'text':
                    item_data['texts'] += 1
                    # Text always has content (from structure)
                    if config.get('full_text') or config.get('preview'):
                        has_data = True
                        item_data['texts_ok'] += 1
                
                elif comp.component_type == 'table':
                    item_data['tables'] += 1
                    if config.get('extracted_data') and config['extracted_data'].get('data'):
                        has_data = True
                        item_data['tables_ok'] += 1
                
                elif comp.component_type == 'chart':
                    item_data['charts'] += 1
                    cd = config.get('chart_data', {})
                    if cd.get('series') or cd.get('type') == 'image':
                        has_data = True
                        item_data['charts_ok'] += 1
                
                item_data['components'].append({
                    'ref_id': comp.ref_id,
                    'type': comp.component_type,
                    'title': comp.title[:30] if comp.title else '',
                    'has_data': has_data
                })
            
            axis_data['items'].append(item_data)
            axis_data['totals']['texts'] += item_data['texts']
            axis_data['totals']['texts_ok'] += item_data['texts_ok']
            axis_data['totals']['tables'] += item_data['tables']
            axis_data['totals']['tables_ok'] += item_data['tables_ok']
            axis_data['totals']['charts'] += item_data['charts']
            axis_data['totals']['charts_ok'] += item_data['charts_ok']
            
            checklist['summary']['total_items'] += 1
            checklist['summary']['total_components'] += len(item_data['components'])
        
        checklist['axes'].append(axis_data)
        checklist['summary']['texts']['total'] += axis_data['totals']['texts']
        checklist['summary']['texts']['with_content'] += axis_data['totals']['texts_ok']
        checklist['summary']['tables']['total'] += axis_data['totals']['tables']
        checklist['summary']['tables']['with_data'] += axis_data['totals']['tables_ok']
        checklist['summary']['charts']['total'] += axis_data['totals']['charts']
        checklist['summary']['charts']['with_data'] += axis_data['totals']['charts_ok']
    
    return checklist


def print_checklist(checklist):
    """Print checklist in readable format."""
    s = checklist['summary']
    
    print("\n" + "=" * 80)
    print("📊 ملخص التقدم الإجمالي")
    print("=" * 80)
    
    print(f"\n📁 البنود: {s['total_items']}")
    print(f"📦 المكونات: {s['total_components']}")
    
    print(f"\n📝 النصوص: {s['texts']['with_content']}/{s['texts']['total']} ", end="")
    print(f"({100*s['texts']['with_content']//s['texts']['total'] if s['texts']['total'] else 0}%)")
    
    print(f"📊 الجداول: {s['tables']['with_data']}/{s['tables']['total']} ", end="")
    print(f"({100*s['tables']['with_data']//s['tables']['total'] if s['tables']['total'] else 0}%)")
    
    print(f"📈 الرسوم: {s['charts']['with_data']}/{s['charts']['total']} ", end="")
    print(f"({100*s['charts']['with_data']//s['charts']['total'] if s['charts']['total'] else 0}%)")
    
    print("\n" + "=" * 80)
    print("📋 التفصيل حسب المحاور")
    print("=" * 80)
    
    for axis in checklist['axes']:
        t = axis['totals']
        print(f"\n{'─'*80}")
        print(f"📁 المحور {axis['code']}: {axis['name']}")
        print(f"{'─'*80}")
        print(f"{'البند':<8} {'📝 نص':<12} {'📊 جدول':<12} {'📈 شكل':<12} {'الحالة':<10}")
        print(f"{'─'*80}")
        
        for item in axis['items']:
            texts = f"{item['texts_ok']}/{item['texts']}"
            tables = f"{item['tables_ok']}/{item['tables']}"
            charts = f"{item['charts_ok']}/{item['charts']}"
            
            all_ok = (item['texts_ok'] == item['texts'] and 
                     item['tables_ok'] == item['tables'] and 
                     item['charts_ok'] == item['charts'])
            
            status = "✅" if all_ok else "⏳"
            if item['texts'] + item['tables'] + item['charts'] == 0:
                status = "⚪"  # Empty item
            
            print(f"{item['code']:<8} {texts:<12} {tables:<12} {charts:<12} {status:<10}")
        
        # Axis totals
        print(f"{'─'*80}")
        texts_pct = 100*t['texts_ok']//t['texts'] if t['texts'] else 0
        tables_pct = 100*t['tables_ok']//t['tables'] if t['tables'] else 0
        charts_pct = 100*t['charts_ok']//t['charts'] if t['charts'] else 0
        print(f"{'المجموع':<8} {t['texts_ok']}/{t['texts']} ({texts_pct}%) ", end="")
        print(f"{t['tables_ok']}/{t['tables']} ({tables_pct}%) ", end="")
        print(f"{t['charts_ok']}/{t['charts']} ({charts_pct}%)")


def save_checklist_excel(checklist, template):
    """Save checklist to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress Checklist"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    
    row = 1
    
    # Headers
    headers = ['المحور', 'البند', 'العنوان', '📝 نصوص', '📊 جداول', '📈 رسوم', 'الحالة']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    
    # Data
    for axis in checklist['axes']:
        for item in axis['items']:
            ws.cell(row=row, column=1, value=axis['code']).border = border
            ws.cell(row=row, column=2, value=item['code']).border = border
            ws.cell(row=row, column=3, value=item['name']).border = border
            ws.cell(row=row, column=4, value=f"{item['texts_ok']}/{item['texts']}").border = border
            ws.cell(row=row, column=5, value=f"{item['tables_ok']}/{item['tables']}").border = border
            ws.cell(row=row, column=6, value=f"{item['charts_ok']}/{item['charts']}").border = border
            
            all_ok = (item['texts_ok'] == item['texts'] and 
                     item['tables_ok'] == item['tables'] and 
                     item['charts_ok'] == item['charts'])
            
            total = item['texts'] + item['tables'] + item['charts']
            
            if total == 0:
                status = "⚪ فارغ"
                fill = empty_fill
            elif all_ok:
                status = "✅ مكتمل"
                fill = ok_fill
            else:
                status = "⏳ جزئي"
                fill = warn_fill
            
            cell = ws.cell(row=row, column=7, value=status)
            cell.border = border
            cell.fill = fill
            
            row += 1
    
    # Summary at end
    row += 1
    s = checklist['summary']
    ws.cell(row=row, column=1, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"{s['texts']['with_content']}/{s['texts']['total']}")
    ws.cell(row=row, column=5, value=f"{s['tables']['with_data']}/{s['tables']['total']}")
    ws.cell(row=row, column=6, value=f"{s['charts']['with_data']}/{s['charts']['total']}")
    
    output_path = os.path.join(DATA_DIR, 'progress_checklist.xlsx')
    wb.save(output_path)
    
    # Copy to desktop
    import shutil
    shutil.copy(output_path, os.path.expanduser('~/Desktop/progress_checklist.xlsx'))
    
    return output_path


def main():
    print("=" * 80)
    print("🔄 إكمال ربط جميع البيانات")
    print("=" * 80)
    
    # Load data
    tables, charts = load_extracted_data()
    print(f"\n📊 الجداول المستخرجة: {len(tables)}")
    print(f"📈 الرسوم المستخرجة: {len(charts)}")
    
    # Get template
    template = Template.objects.get(id=1)
    
    # Link tables
    tables_linked = link_tables(tables, template)
    
    # Link charts
    charts_linked = link_charts(charts, template)
    
    # Generate checklist
    checklist = generate_checklist(template)
    
    # Print checklist
    print_checklist(checklist)
    
    # Save to Excel
    excel_path = save_checklist_excel(checklist, template)
    print(f"\n💾 Excel saved: {excel_path}")
    print(f"💾 Also copied to: ~/Desktop/progress_checklist.xlsx")
    
    # Save JSON
    json_path = os.path.join(DATA_DIR, 'progress_checklist.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(checklist, f, ensure_ascii=False, indent=2)
    print(f"💾 JSON saved: {json_path}")


if __name__ == '__main__':
    main()
